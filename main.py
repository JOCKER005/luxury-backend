from fastapi import FastAPI, HTTPException, Depends, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from contextlib import asynccontextmanager
from collections import defaultdict
import os, io, time, json, hashlib
import httpx
import hmac as hmac_lib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timezone, timedelta
from typing import Optional
import mercadopago

from database import engine, Base, DATABASE_URL
from models import Product, Order, OrderItem
from schemas import (
    ProductCreate, ProductUpdate, ProductOut,
    OrderOut, PaymentPreferenceRequest, PaymentPreferenceResponse
)
from sqlalchemy.orm import Session
from database import get_db

# ─── Engine detection ─────────────────────────────────────────────────────────
IS_POSTGRES = DATABASE_URL.startswith("postgresql")

def lock_row(query):
    return query.with_for_update() if IS_POSTGRES else query

# ─── Brute-force protection ───────────────────────────────────────────────────
_failed: dict = defaultdict(list)
MAX_ATTEMPTS = 5
WINDOW   = 300   # 5 min
LOCKOUT  = 600   # 10 min

def check_brute_force(ip: str):
    now = time.time()
    _failed[ip] = [t for t in _failed[ip] if now - t < WINDOW]
    if len(_failed[ip]) >= MAX_ATTEMPTS:
        wait = int(LOCKOUT - (now - min(_failed[ip])))
        raise HTTPException(429, detail=f"Demasiados intentos. Esperá {wait//60}m {wait%60}s.")

def record_fail(ip: str):   _failed[ip].append(time.time())
def clear_attempts(ip: str): _failed.pop(ip, None)

# ─── Lifespan ────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    Base.metadata.create_all(bind=engine)
    try:
        db = next(get_db())
        if db.query(Product).count() == 0:
            seed_products(db)
        db.close()
    except Exception as e:
        print(f"[WARN] Seed falló: {e}")
    yield

app = FastAPI(title="Luxe Joyería API", version="1.0.0", lifespan=lifespan)

# ─── CORS ─────────────────────────────────────────────────────────────────────
_raw = os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:5173,http://localhost:4173,https://luxury-front-five.vercel.app"
)
ALLOWED_ORIGINS = [o.strip() for o in _raw.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,  # o ["*"] para debug
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Auth ─────────────────────────────────────────────────────────────────────
ADMIN_SECRET = os.getenv("ADMIN_SECRET", "luxe-admin-secret-2024")

def verify_admin(request: Request, x_admin_token: str = Header(...)):
    ip = request.client.host if request.client else "unknown"
    check_brute_force(ip)
    if x_admin_token != ADMIN_SECRET:
        record_fail(ip)
        raise HTTPException(401, detail="Token de admin inválido")
    clear_attempts(ip)
    return True

# ─── MP SDK ───────────────────────────────────────────────────────────────────
def get_mp_sdk():
    token = os.getenv("MP_ACCESS_TOKEN")
    if not token:
        raise HTTPException(500, detail="MP_ACCESS_TOKEN no configurado")
    return mercadopago.SDK(token)

# ─── Public: Products ─────────────────────────────────────────────────────────
@app.get("/api/products", response_model=list[ProductOut])
def list_products(featured: Optional[bool] = None, category: Optional[str] = None,
                  db: Session = Depends(get_db)):
    q = db.query(Product).filter(Product.active == True)
    if featured is not None: q = q.filter(Product.featured == featured)
    if category:             q = q.filter(Product.category == category)
    return q.all()


@app.get("/api/products/{product_id}", response_model=ProductOut)
def get_product(product_id: int, db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.id == product_id, Product.active == True).first()
    if not p: raise HTTPException(404, detail="Producto no encontrado")
    return p

# ─── Public: Payments ─────────────────────────────────────────────────────────
@app.post("/api/payments/create-preference", response_model=PaymentPreferenceResponse)
def create_preference(body: PaymentPreferenceRequest, db: Session = Depends(get_db)):
    sdk = get_mp_sdk()

    # FIX: Antes se hacían 2 queries por ítem (una para validar stock, otra para obtener precio).
    # Ahora se hace 1 sola query por ítem y se reutiliza el objeto en ambas fases.
    product_map: dict[int, Product] = {}
    for item in body.items:
        if item.productId not in product_map:
            p = lock_row(
                db.query(Product).filter(Product.id == item.productId, Product.active == True)
            ).first()
            if not p:
                raise HTTPException(404, detail=f"'{item.name}' no está disponible")
            product_map[item.productId] = p

        p = product_map[item.productId]
        if p.stock is not None and p.stock < item.quantity:
            raise HTTPException(409, detail=f"Stock insuficiente para '{p.name}'. Disponible: {p.stock}")

    # Precios SIEMPRE desde la DB — nunca del cliente
    mp_items, total_real = [], 0.0
    for item in body.items:
        p = product_map[item.productId]
        precio = float(p.price)
        total_real += precio * item.quantity
        mp_items.append({
            "id": str(item.productId), "title": p.name,
            "quantity": item.quantity, "unit_price": precio,
            "currency_id": "ARS", "picture_url": item.image,
        })

    result = sdk.preference().create({
        "items": mp_items,
        "back_urls": {
            "success": body.backUrls.success,
            "failure": body.backUrls.failure,
            "pending": body.backUrls.pending,
        },
        "auto_return": "approved",
        "statement_descriptor": "Luxe Joyería",
    })
    pref = result["response"]
    if result["status"] not in (200, 201):
        raise HTTPException(500, detail=str(pref))

    order = Order(mp_preference_id=pref["id"], status="pending",
                  total=total_real, created_at=datetime.now(timezone.utc))
    db.add(order)
    db.flush()

    for item in body.items:
        p = product_map[item.productId]
        db.add(OrderItem(
            order_id=order.id, product_id=item.productId,
            product_name=p.name, quantity=item.quantity, unit_price=float(p.price)
        ))
    db.commit()

    return PaymentPreferenceResponse(
        id=pref["id"],
        initPoint=pref.get("init_point", ""),
        sandboxInitPoint=pref.get("sandbox_init_point", ""),
    )


@app.post("/api/payments/webhook")
async def mp_webhook(request: Request, db: Session = Depends(get_db)):
    raw_body = await request.body()  # leer UNA sola vez

    mp_secret = os.getenv("MP_WEBHOOK_SECRET")
    if mp_secret:
        x_sig = request.headers.get("x-signature", "")
        x_rid = request.headers.get("x-request-id", "")
        try:
            parts = dict(p.split("=", 1) for p in x_sig.split(";") if "=" in p)
            ts, v1 = parts.get("ts", ""), parts.get("v1", "")
            if not ts or not v1:
                raise HTTPException(401, detail="Firma incompleta")
            manifest = f"id:{x_rid};request-id:{x_rid};ts:{ts};"
            expected = hmac_lib.new(mp_secret.encode(), manifest.encode(), hashlib.sha256).hexdigest()
            if not hmac_lib.compare_digest(expected, v1):
                raise HTTPException(401, detail="Firma inválida")
        except HTTPException: raise
        except Exception: raise HTTPException(400, detail="Error verificando firma")

    try:
        body = json.loads(raw_body)
    except Exception:
        raise HTTPException(400, detail="Body inválido")

    if body.get("type") == "payment":
        payment_id = body.get("data", {}).get("id")
        if payment_id:
            result  = get_mp_sdk().payment().get(payment_id)
            payment = result["response"]
            pref_id, status = payment.get("preference_id"), payment.get("status")

            order = db.query(Order).filter(Order.mp_preference_id == pref_id).first()
            if not order:
                print(f"[WARN] Webhook sin orden para preference_id={pref_id}")
                return {"ok": True, "msg": "order_not_found"}

            if order.mp_payment_id == str(payment_id) and order.status == "approved":
                return {"ok": True, "msg": "already_processed"}

            order.status       = status
            order.mp_payment_id = str(payment_id)
            db.commit()

            if status == "approved":
                for item in db.query(OrderItem).filter(OrderItem.order_id == order.id).all():
                    p = lock_row(db.query(Product).filter(Product.id == item.product_id)).first()
                    if p and p.stock is not None:
                        p.stock = max(0, p.stock - item.quantity)
                db.commit()

    return {"ok": True}

# ─── Admin: Products ──────────────────────────────────────────────────────────
@app.get("/api/admin/products", response_model=list[ProductOut])
def admin_list_products(db: Session = Depends(get_db), _=Depends(verify_admin)):
    return db.query(Product).order_by(Product.id.desc()).all()


@app.post("/api/admin/products", response_model=ProductOut)
def create_product(body: ProductCreate, db: Session = Depends(get_db), _=Depends(verify_admin)):
    data = body.model_dump()
    # Serializar images list a JSON string para guardar en DB
    data["images"] = json.dumps(data.get("images") or [])
    p = Product(**data)
    db.add(p); db.commit(); db.refresh(p)
    return p


@app.patch("/api/admin/products/{product_id}", response_model=ProductOut)
def update_product(product_id: int, body: ProductUpdate,
                   db: Session = Depends(get_db), _=Depends(verify_admin)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p: raise HTTPException(404, detail="Producto no encontrado")
    data = body.model_dump(exclude_unset=True)
    # Serializar images list a JSON string si viene en el body
    if "images" in data:
        data["images"] = json.dumps(data["images"] or [])
    for field, val in data.items():
        setattr(p, field, val)
    db.commit(); db.refresh(p)
    return p


@app.delete("/api/admin/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), _=Depends(verify_admin)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p: raise HTTPException(404, detail="Producto no encontrado")
    pending = (
        db.query(OrderItem).join(Order)
        .filter(OrderItem.product_id == product_id, Order.status.in_(["pending", "in_process"]))
        .count()
    )
    if pending > 0:
        raise HTTPException(409, detail=f"No se puede eliminar: tiene {pending} orden(es) pendiente(s).")
    p.active = False; db.commit()
    return {"ok": True}

# ─── Admin: Orders ────────────────────────────────────────────────────────────
@app.get("/api/admin/orders", response_model=list[OrderOut])
def admin_list_orders(status: Optional[str] = None,
                      db: Session = Depends(get_db), _=Depends(verify_admin)):
    q = db.query(Order)
    if status: q = q.filter(Order.status == status)
    return q.order_by(Order.created_at.desc()).all()


@app.patch("/api/admin/orders/{order_id}/status")
def update_order_status(order_id: int, body: dict,
                        db: Session = Depends(get_db), _=Depends(verify_admin)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order: raise HTTPException(404, detail="Orden no encontrada")
    order.status = body.get("status", order.status); db.commit()
    return {"ok": True}

# ─── Admin: Excel ─────────────────────────────────────────────────────────────
@app.get("/api/admin/export/libro-diario")
def export_libro_diario(date_from: Optional[str] = None, date_to: Optional[str] = None,
                        db: Session = Depends(get_db), _=Depends(verify_admin)):
    q = db.query(Order)
    if date_from: q = q.filter(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:   q = q.filter(Order.created_at <= datetime.fromisoformat(date_to))
    orders = q.order_by(Order.created_at.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Diario de Ventas"
    gold, dark, light_gold = "C9A227", "1A1A1A", "F5E6C3"
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "LUXE JOYERÍA — Libro Diario de Ventas"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color=gold)
    ws["A1"].alignment = center
    ws["A1"].fill      = PatternFill("solid", fgColor="0D0D0D")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font      = Font(name="Calibri", size=10, color="888888", italic=True)
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 20

    headers = ["# Orden","Fecha (AR)","Estado","MP Preference ID","MP Payment ID","Productos","Cantidad","Total (ARS)"]
    ws.row_dimensions[3].height = 28
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=dark)
        cell.alignment = center
        cell.border    = border

    total_general = 0.0
    if not orders:
        ws.merge_cells("A4:H4")
        c = ws.cell(row=4, column=1, value="No hay órdenes en el período seleccionado")
        c.font      = Font(name="Calibri", italic=True, color="888888")
        c.alignment = center
        total_row = 5
    else:
        for i, order in enumerate(orders, 4):
            items  = db.query(OrderItem).filter(OrderItem.order_id == order.id).all()
            prods  = ", ".join(f"{it.product_name} x{it.quantity}" for it in items)
            cant   = sum(it.quantity for it in items)
            fill   = PatternFill("solid", fgColor=light_gold) if i % 2 == 0 \
                     else PatternFill("solid", fgColor="FFFFFF")
            fecha  = ""
            if order.created_at:
                dt_ar = order.created_at - timedelta(hours=3)
                fecha = dt_ar.strftime("%d/%m/%Y %H:%M") + " AR"

            row_data = [order.id, fecha, order.status.upper() if order.status else "",
                        order.mp_preference_id or "", order.mp_payment_id or "",
                        prods, cant, order.total or 0]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col == 8: cell.number_format = '"$"#,##0.00'
                if col == 3:
                    color = ("1A7F37" if order.status == "approved"
                             else "CF1124" if order.status == "rejected" else "B45309")
                    cell.font = Font(name="Calibri", bold=True, color=color)
            total_general += order.total or 0
        total_row = len(orders) + 4

    ws.merge_cells(f"A{total_row}:G{total_row}")
    tl = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
    tl.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    tl.fill = PatternFill("solid", fgColor=dark)
    tl.alignment = Alignment(horizontal="right", vertical="center")
    tv = ws.cell(row=total_row, column=8, value=total_general)
    tv.font = Font(name="Calibri", bold=True, size=12, color=dark)
    tv.fill = PatternFill("solid", fgColor=gold)
    tv.number_format = '"$"#,##0.00'
    tv.alignment = center
    ws.row_dimensions[total_row].height = 32

    for i, w in enumerate([10,22,14,36,22,50,12,16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"libro_diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

# ─── Public: Market Prices ────────────────────────────────────────────────────
@app.get("/api/market/prices")
async def get_market_prices():
    """
    Consulta precio del oro desde Binance (XAUUSDT) y dólar desde bluelytics.
    Se hace desde el backend para evitar bloqueos CORS en el navegador.
    """
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            gold_res  = await client.get("https://api.binance.com/api/v3/ticker/price?symbol=XAUUSDT")
            dolar_res = await client.get("https://api.bluelytics.com.ar/v2/latest")

        gold_data  = gold_res.json()
        dolar_data = dolar_res.json()

        gold_usd  = float(gold_data["price"]) if "price" in gold_data else None
        blue      = dolar_data.get("blue", {}).get("value_sell")
        oficial   = dolar_data.get("oficial", {}).get("value_sell")

        return {
            "goldUSD":      gold_usd,
            "dolarBlue":    blue,
            "dolarOficial": oficial,
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"No se pudo obtener precios: {e}")

# ─── Seed ─────────────────────────────────────────────────────────────────────
def seed_products(db: Session):
    db.add_all([
        Product(name="Anillo Eternity Diamantes",  category="anillos",  price=45000, stock=5,  featured=True,  active=True, rating=4.9, reviews=127,
                description="Anillo de eternidad con diamantes engastados en oro blanco 18k.",
                image="https://images.unsplash.com/photo-1605100804763-247f67b2548e?auto=format&fit=crop&w=600&q=80"),
        Product(name="Collar Cadena Veneciana Oro", category="collares", price=28000, stock=8,  featured=True,  active=True, rating=4.8, reviews=89,
                description="Collar de cadena veneciana en oro amarillo 18k.",
                image="https://images.unsplash.com/photo-1611591437281-460bfbe1220a?auto=format&fit=crop&w=600&q=80"),
        Product(name="Aretes Gota Zafiro Azul",    category="aretes",   price=32000, stock=4,  featured=True,  active=True, rating=4.9, reviews=54,
                description="Aretes de gota con zafiros azules naturales rodeados de diamantes.",
                image="https://images.unsplash.com/photo-1535632066927-ab7c9ab60908?auto=format&fit=crop&w=600&q=80"),
        Product(name="Pulsera Tennis Diamantes",   category="pulseras", price=62000, stock=3,  featured=True,  active=True, rating=5.0, reviews=38,
                description="Pulsera tennis con diamantes de 0.05ct en oro blanco 18k.",
                image="https://images.unsplash.com/photo-1573408301185-9519f94816ec?auto=format&fit=crop&w=600&q=80"),
        Product(name="Anillo Solitario Brillante", category="anillos",  price=89000, stock=2,  featured=False, active=True, rating=5.0, reviews=21,
                description="Anillo solitario con diamante central de 0.5ct en oro rosado 18k.",
                image="https://images.unsplash.com/photo-1518131672697-613becd4fab5?auto=format&fit=crop&w=600&q=80"),
        Product(name="Collar Perlas Cultivadas",   category="collares", price=19500, stock=10, featured=False, active=True, rating=4.7, reviews=63,
                description="Collar de perlas de agua dulce cultivadas, cierre en plata 925.",
                image="https://images.unsplash.com/photo-1599643477874-c4ea90b50369?auto=format&fit=crop&w=600&q=80"),
    ])
    db.commit()
